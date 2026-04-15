#!/usr/bin/env python3
"""
mSCOA v6.9 Chart Seeder - Direct DB insertion
Extracts data from GrapMapping_V6.9.xlsx and loads directly into PostgreSQL.
"""
import openpyxl
import json
import os
import sys
import uuid
import psycopg2
from datetime import datetime

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'docs', 'mscoa', 'GrapMapping_V6.9.xlsx')
DB_URL = os.environ.get('DATABASE_URL', '')

def get_connection():
    return psycopg2.connect(DB_URL)

def create_version(conn):
    cur = conn.cursor()
    version_id = str(uuid.uuid4())

    cur.execute("SELECT id FROM mscoa_chart_versions WHERE version = '6.9' LIMIT 1")
    row = cur.fetchone()
    if row:
        version_id = row[0]
        print(f'  Existing version found: {version_id}, clearing old data...')
        cur.execute("DELETE FROM mscoa_chart_items WHERE \"chartVersionId\" = %s", (version_id,))
        cur.execute("DELETE FROM mscoa_grap_mappings WHERE \"chartVersionId\" = %s", (version_id,))
        conn.commit()
        return version_id

    cur.execute("""
        INSERT INTO mscoa_chart_versions (id, version, label, description, "sourceFileName", "effectiveFrom", "effectiveTo", status, "isCurrent", "totalAccounts", "postingLevelAccounts", "createdAt", "updatedAt")
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        version_id, '6.9', 'mSCOA v6.9',
        'Municipal Standard Chart of Accounts version 6.9 - current version effective until 30 June 2026',
        'GrapMapping_V6.9.xlsx',
        '2024-07-01', '2026-06-30',
        'draft', False, 0, 0,
        datetime.utcnow(), datetime.utcnow()
    ))
    conn.commit()
    print(f'  Created version: {version_id}')
    return version_id

def load_chart_items(conn, version_id):
    print(f'Opening {XLSX_PATH}...')
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['6.9']

    cur = conn.cursor()
    batch = []
    total = 0
    posting_count = 0
    segment_stats = {}

    print('Extracting chart items from 6.9 sheet...')
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=210000, max_col=30, values_only=True)):
        acct = str(row[1]) if row[1] else ''
        segment = str(row[2]) if row[2] else ''
        if not acct or not segment:
            continue

        is_posting = str(row[28]) == 'Y' if row[28] else False
        scoa_file = str(row[6]) if row[6] else ''
        level = int(row[8]) if row[8] else 1

        def clean(val, maxlen=500):
            if val is None:
                return None
            s = str(val)
            return s[:maxlen] if s != 'None' else None

        item_id = str(uuid.uuid4())
        batch.append((
            item_id, version_id, acct, segment,
            clean(row[3]), clean(row[4]),
            scoa_file, clean(row[7], 500), clean(row[23], 200),
            level, is_posting,
            clean(row[9], 100), clean(row[29], 10),
            clean(row[10], 200), clean(row[11], 200), clean(row[12], 200),
            clean(row[13], 200), clean(row[14], 200), clean(row[15], 200),
            clean(row[16], 200), clean(row[17], 200), clean(row[18], 200),
            clean(row[19], 200), clean(row[20], 200), clean(row[21], 200),
            clean(row[24], 20), clean(row[25], 5), clean(row[26], 50),
            clean(row[27], 50), clean(row[5], 500),
            datetime.utcnow()
        ))

        if is_posting:
            posting_count += 1

        if segment not in segment_stats:
            segment_stats[segment] = {'scoaFile': scoa_file, 'total': 0, 'posting': 0}
        segment_stats[segment]['total'] += 1
        if is_posting:
            segment_stats[segment]['posting'] += 1

        total += 1

        if len(batch) >= 1000:
            insert_chart_batch(cur, batch)
            batch = []
            if total % 10000 == 0:
                print(f'  Loaded {total} chart items...')
                conn.commit()

    if batch:
        insert_chart_batch(cur, batch)

    conn.commit()
    wb.close()

    cur.execute("""
        UPDATE mscoa_chart_versions
        SET "totalAccounts" = %s, "postingLevelAccounts" = %s, "segmentSummary" = %s, "updatedAt" = %s
        WHERE id = %s
    """, (total, posting_count, json.dumps(segment_stats), datetime.utcnow(), version_id))
    conn.commit()

    print(f'  Total chart items: {total}')
    print(f'  Posting level items: {posting_count}')
    print(f'  Segments: {", ".join(sorted(segment_stats.keys()))}')
    return total

def insert_chart_batch(cur, batch):
    args_str = ','.join(cur.mogrify(
        "(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", b
    ).decode('utf-8') for b in batch)
    cur.execute(f"""
        INSERT INTO mscoa_chart_items (
            id, "chartVersionId", "accountNumber", segment,
            "scoaGuid", "parentScoaGuid",
            "scoaFile", description, "shortDescription",
            level, "postingLevel",
            "statementType", "accountPrefix",
            l1, l2, l3, l4, l5, l6, l7, l8, l9, l10, l11, l12,
            "vatStatus", "breakdownAllowed", principle, "applicableTo",
            "definitionDescription", "createdAt"
        ) VALUES {args_str}
    """)

def load_grap_mappings(conn, version_id):
    print(f'Opening {XLSX_PATH} for GRAP mappings...')
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['SOFP MAPPING']

    cur = conn.cursor()
    batch = []
    total = 0

    print('Extracting GRAP mappings from SOFP MAPPING sheet...')
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=155000, max_col=15, values_only=True)):
        acct = str(row[1]) if row[1] else ''
        if not acct:
            continue

        def clean(val, maxlen=500):
            if val is None:
                return None
            s = str(val)
            return s[:maxlen] if s != 'None' else None

        is_posting = str(row[5]) == 'Y' if row[5] else False
        item_id = str(uuid.uuid4())

        batch.append((
            item_id, version_id,
            clean(row[0], 50), acct,
            clean(row[2], 500), clean(row[3], 10),
            clean(row[4]), is_posting,
            clean(row[6], 100), clean(row[7], 100),
            clean(row[8], 200), clean(row[9], 200),
            clean(row[10], 200), clean(row[11], 200),
            datetime.utcnow()
        ))

        total += 1

        if len(batch) >= 1000:
            insert_grap_batch(cur, batch)
            batch = []
            if total % 10000 == 0:
                print(f'  Loaded {total} GRAP mappings...')
                conn.commit()

    if batch:
        insert_grap_batch(cur, batch)

    conn.commit()
    wb.close()
    print(f'  Total GRAP mappings: {total}')
    return total

def insert_grap_batch(cur, batch):
    args_str = ','.join(cur.mogrify(
        "(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", b
    ).decode('utf-8') for b in batch)
    cur.execute(f"""
        INSERT INTO mscoa_grap_mappings (
            id, "chartVersionId",
            "sourceId", "accountNumber",
            "scoaDescription", prefix,
            "scoaGuid", "postingLevel",
            "statementType", category,
            "reportingItem", "subClass",
            "subClassBreakdown", "subClassBreakdownDetail",
            "createdAt"
        ) VALUES {args_str}
    """)

def activate_version(conn, version_id):
    cur = conn.cursor()
    cur.execute("UPDATE mscoa_chart_versions SET \"isCurrent\" = false")
    cur.execute("UPDATE mscoa_chart_versions SET \"isCurrent\" = true, status = 'active', \"updatedAt\" = %s WHERE id = %s",
                (datetime.utcnow(), version_id))
    conn.commit()
    print('  Version activated!')

def main():
    print('=== mSCOA v6.9 Chart Seeder (Direct DB) ===')
    print()

    if not os.path.exists(XLSX_PATH):
        print(f'ERROR: Excel file not found at {XLSX_PATH}')
        sys.exit(1)

    if not DB_URL:
        print('ERROR: DATABASE_URL not set')
        sys.exit(1)

    conn = get_connection()
    print('Connected to database')

    version_id = create_version(conn)
    chart_count = load_chart_items(conn, version_id)
    grap_count = load_grap_mappings(conn, version_id)
    activate_version(conn, version_id)

    conn.close()

    print()
    print('=== Seeding Complete ===')
    print(f'  Chart items: {chart_count}')
    print(f'  GRAP mappings: {grap_count}')
    print(f'  Version ID: {version_id}')
    print(f'  Status: ACTIVE')

if __name__ == '__main__':
    main()
