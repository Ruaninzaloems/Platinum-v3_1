#!/usr/bin/env python3
"""
mSCOA v6.9 Chart Seeder
Extracts data from GrapMapping_V6.9.xlsx and loads into PostgreSQL via the NestJS API.
Usage: python3 scripts/seed-mscoa.py
"""
import openpyxl
import json
import urllib.request
import sys
import os

API_BASE = 'http://localhost:3000/api'
XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'docs', 'mscoa', 'GrapMapping_V6.9.xlsx')

def api_call(method, path, data=None, token=None):
    url = f'{API_BASE}{path}'
    headers = {'Content-Type': 'application/json'}
    if token:
        headers['Authorization'] = f'Bearer {token}'
    body = json.dumps(data).encode('utf-8') if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=300) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8') if e.fp else ''
        print(f'  API Error {e.code}: {error_body[:200]}')
        raise

def login():
    email = os.environ.get('ADMIN_EMAIL', 'admin@platinum.gov.za')
    password = os.environ.get('ADMIN_PASSWORD')
    if not password:
        print('Error: ADMIN_PASSWORD environment variable is required')
        sys.exit(1)
    print(f'Logging in as {email}...')
    result = api_call('POST', '/auth/login', {
        'email': email,
        'password': password
    })
    return result['accessToken']

def create_version(token):
    print('Creating mSCOA v6.9 chart version...')
    try:
        result = api_call('POST', '/mscoa/versions', {
            'version': '6.9',
            'label': 'mSCOA v6.9',
            'description': 'Municipal Standard Chart of Accounts version 6.9 - current version effective until 30 June 2026',
            'sourceFileName': 'GrapMapping_V6.9.xlsx',
            'effectiveFrom': '2024-07-01',
            'effectiveTo': '2026-06-30',
        }, token)
        print(f'  Created version: {result["id"]}')
        return result['id']
    except Exception as e:
        print(f'  Version may already exist, fetching...')
        versions = api_call('GET', '/mscoa/versions', token=token)
        for v in versions:
            if v['version'] == '6.9':
                print(f'  Found existing version: {v["id"]}')
                return v['id']
        raise

def extract_and_load_chart(version_id, token):
    print(f'Opening {XLSX_PATH}...')
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['6.9']

    print('Extracting chart items from 6.9 sheet...')
    batch = []
    total_loaded = 0
    batch_size = 400

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=210000, max_col=30, values_only=True)):
        acct = str(row[1]) if row[1] else ''
        if not acct:
            continue

        segment = str(row[2]) if row[2] else ''
        if not segment:
            continue

        item = {
            'accountNumber': acct,
            'segment': segment,
            'scoaGuid': str(row[3]) if row[3] else None,
            'parentScoaGuid': str(row[4]) if row[4] else None,
            'scoaFile': str(row[6]) if row[6] else '',
            'description': str(row[7])[:500] if row[7] else '',
            'shortDescription': str(row[23])[:200] if row[23] else '',
            'level': int(row[8]) if row[8] else 1,
            'postingLevel': str(row[28]) == 'Y' if row[28] else False,
            'statementType': str(row[9]) if row[9] else None,
            'accountPrefix': str(row[29]) if row[29] else '',
            'l1': str(row[10])[:200] if row[10] else None,
            'l2': str(row[11])[:200] if row[11] else None,
            'l3': str(row[12])[:200] if row[12] else None,
            'l4': str(row[13])[:200] if row[13] else None,
            'l5': str(row[14])[:200] if row[14] else None,
            'l6': str(row[15])[:200] if row[15] else None,
            'l7': str(row[16])[:200] if row[16] else None,
            'l8': str(row[17])[:200] if row[17] else None,
            'l9': str(row[18])[:200] if row[18] else None,
            'l10': str(row[19])[:200] if row[19] else None,
            'l11': str(row[20])[:200] if row[20] else None,
            'l12': str(row[21])[:200] if row[21] else None,
            'vatStatus': str(row[24])[:20] if row[24] else None,
            'breakdownAllowed': str(row[25])[:5] if row[25] else None,
            'principle': str(row[26])[:50] if row[26] else None,
            'applicableTo': str(row[27])[:50] if row[27] else None,
            'definitionDescription': str(row[5])[:500] if row[5] else None,
        }

        # Clean None strings
        for key in list(item.keys()):
            if item[key] == 'None':
                item[key] = None

        batch.append(item)

        if len(batch) >= batch_size:
            api_call('POST', f'/mscoa/versions/{version_id}/load-chart', {'items': batch}, token)
            total_loaded += len(batch)
            if total_loaded % 5000 == 0:
                print(f'  Loaded {total_loaded} chart items...')
            batch = []

    if batch:
        api_call('POST', f'/mscoa/versions/{version_id}/load-chart', {'items': batch}, token)
        total_loaded += len(batch)

    wb.close()
    print(f'  Total chart items loaded: {total_loaded}')
    return total_loaded

def extract_and_load_grap_mappings(version_id, token):
    print(f'Opening {XLSX_PATH} for GRAP mappings...')
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['SOFP MAPPING']

    print('Extracting GRAP mappings from SOFP MAPPING sheet...')
    batch = []
    total_loaded = 0
    batch_size = 400

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=155000, max_col=15, values_only=True)):
        acct = str(row[1]) if row[1] else ''
        if not acct:
            continue

        mapping = {
            'id': str(row[0]) if row[0] else None,
            'accountNumber': acct,
            'description': str(row[2])[:500] if row[2] else '',
            'prefix': str(row[3])[:10] if row[3] else None,
            'guid': str(row[4]) if row[4] else None,
            'postingLevel': str(row[5]) == 'Y' if row[5] else False,
            'statementType': str(row[6]) if row[6] else None,
            'category': str(row[7]) if row[7] else None,
            'reportingItem': str(row[8]) if row[8] else None,
            'subClass': str(row[9]) if row[9] else None,
            'subClassBreakdown': str(row[10]) if row[10] else None,
            'subClassBreakdownDetail': str(row[11]) if row[11] else None,
        }

        for key in list(mapping.keys()):
            if mapping[key] == 'None':
                mapping[key] = None

        batch.append(mapping)

        if len(batch) >= batch_size:
            api_call('POST', f'/mscoa/versions/{version_id}/load-grap-mappings', {'mappings': batch}, token)
            total_loaded += len(batch)
            if total_loaded % 5000 == 0:
                print(f'  Loaded {total_loaded} GRAP mappings...')
            batch = []

    if batch:
        api_call('POST', f'/mscoa/versions/{version_id}/load-grap-mappings', {'mappings': batch}, token)
        total_loaded += len(batch)

    wb.close()
    print(f'  Total GRAP mappings loaded: {total_loaded}')
    return total_loaded

def activate_version(version_id, token):
    print('Activating version...')
    api_call('POST', f'/mscoa/versions/{version_id}/activate', {}, token)
    print('  Version activated successfully!')

def main():
    print('=== mSCOA v6.9 Chart Seeder ===')
    print()

    if not os.path.exists(XLSX_PATH):
        print(f'ERROR: Excel file not found at {XLSX_PATH}')
        sys.exit(1)

    token = login()
    version_id = create_version(token)
    
    chart_count = extract_and_load_chart(version_id, token)
    grap_count = extract_and_load_grap_mappings(version_id, token)
    
    activate_version(version_id, token)

    print()
    print('=== Seeding Complete ===')
    print(f'  Chart items: {chart_count}')
    print(f'  GRAP mappings: {grap_count}')
    print(f'  Version ID: {version_id}')

if __name__ == '__main__':
    main()
